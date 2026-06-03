#!/usr/bin/env python3
"""
sync-excel.py — Version11_(2026).xlsx 에서 정도검사 판정 기준값 추출
출력: src/precision-criteria.json

사용:
  python3 scripts/sync-excel.py
  (또는 npm run sync-excel)
"""

import json, re, sys, os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)

# ── 경로 설정 ─────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
EXCEL_PATH  = os.path.join(os.path.dirname(PROJECT_DIR), "Version11_(2026).xlsx")
OUTPUT_PATH = os.path.join(PROJECT_DIR, "src", "precision-criteria.json")

if not os.path.exists(EXCEL_PATH):
    # 현재 디렉토리에서도 검색
    alt = os.path.join(os.path.expanduser("~"), "coding", "Version11_(2026).xlsx")
    if os.path.exists(alt):
        EXCEL_PATH = alt
    else:
        print(f"[ERROR] 엑셀 파일을 찾을 수 없습니다:\n  {EXCEL_PATH}\n  {alt}")
        sys.exit(1)

print(f"[sync-excel] 엑셀 읽는 중: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
ws = wb["Sheet1"]

# ── 수식에서 숫자 추출 헬퍼 ──────────────────────────────
def extract_limit(formula, patterns):
    """정규식 패턴 목록으로 한계값 추출"""
    if not formula:
        return None
    for pat in patterns:
        m = re.search(pat, formula)
        if m:
            return float(m.group(1))
    return None

def extract_round_digits(formula):
    """ROUND(expr, n) 에서 n 추출"""
    m = re.search(r"ROUND\([^,]+,\s*(\d+)\)", formula, re.IGNORECASE)
    return int(m.group(1)) if m else None

# ── 각 판정 셀에서 기준값 추출 ────────────────────────────
criteria = {}

# 반복성 (F40): D40 > 3 → 부적합
f40 = ws["F40"].value or ""
rep_limit = extract_limit(f40, [r"D\d+\s*>\s*(\d+(?:\.\d+)?)"])
rep_round = extract_round_digits(ws["D40"].value or "")

# 드리프트 (F43, F46): D43 <= 5 → 적합
f43 = ws["F43"].value or ""
f46 = ws["F46"].value or ""
zero_limit = extract_limit(f43, [r"D\d+\s*<=\s*(\d+(?:\.\d+)?)"])
span_limit  = extract_limit(f46, [r"D\d+\s*<=\s*(\d+(?:\.\d+)?)"])
drift_round = extract_round_digits(ws["D43"].value or "")

# 직선성 (F50): AND(D50 >= -5, D50 <= 5)
f50 = ws["F50"].value or ""
lin_limit = extract_limit(f50, [r"D\d+\s*<=\s*(\d+(?:\.\d+)?)"])
lin_round  = extract_round_digits(ws["D50"].value or "")

# 응답시간 (F51): D24 > 15 → 부적합
f51 = ws["F51"].value or ""
resp_limit = extract_limit(f51, [r"D\d+\s*>\s*(\d+(?:\.\d+)?)"])

# ── 결과 출력 ─────────────────────────────────────────────
criteria = {
    "_source":  os.path.basename(EXCEL_PATH),
    "_synced":  datetime.now().strftime("%Y-%m-%d %H:%M"),
    "_note":    "이 파일은 sync-excel.py가 자동 생성합니다. 직접 수정하지 마세요.",
    "repeatabilityRsd":   rep_limit   if rep_limit  is not None else 3,
    "repeatabilityRound": rep_round   if rep_round  is not None else 1,
    "zeroDrift":          zero_limit  if zero_limit is not None else 5,
    "spanDrift":          span_limit  if span_limit is not None else 5,
    "driftRound":         drift_round if drift_round is not None else 1,
    "linearity":          lin_limit   if lin_limit  is not None else 5,
    "linearityRound":     lin_round   if lin_round  is not None else 1,
    "responseTime":       resp_limit  if resp_limit is not None else 15,
}

os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(criteria, f, ensure_ascii=False, indent=2)

print(f"[sync-excel] 완료 → {OUTPUT_PATH}")
print(json.dumps(criteria, ensure_ascii=False, indent=2))
